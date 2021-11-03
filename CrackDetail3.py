# -*- coding: utf-8 -*-
"""
Created on Mon Nov  1 14:37:03 2021

@author: ssafavizadeh
"""

import cx_Oracle
#import numpy as np
#import pandas as pd
import datetime 
import tkinter as tk
import tkinter.font as font
from tkinter import ttk
from tkinter import Variable
import openpyxl
from datetime import date

try:
    from ctypes import windll
    windll.shcore.SetProcessDpiAwareness(1)
except:
    pass

root = tk.Tk()
root.title("Ad Hoc Crack Detail Data 1.0.0")

#set_dpi_awareness()

font.nametofont("TkDefaultFont").configure(size=12)

#main = ttk.Frame(root, padding=(60, 30))
main = tk.LabelFrame(root, text="Input")
main.pack(padx=10, pady=10, expand="yes")


#result = ttk.Frame(root, padding=(60, 30))
result = tk.LabelFrame(root, text="Result")
result.pack(padx=10, pady=10, expand="yes")

root.columnconfigure(0, weight=1)

connection = cx_Oracle.connect(user="pav_conhist", password="conhist1pav",
                               dsn="ASSET_NEW",
                               encoding="UTF-8")
## Defining Variables ##
selected_hmis_year = tk.StringVar()
selected_data_year = tk.StringVar()
direction = tk.StringVar()
dir2 = tk.StringVar()
selected_county = tk.StringVar()
selected_route = tk.StringVar()
bmp = tk.StringVar()
emp = tk.StringVar()

dataYear = tk.StringVar()
county = tk.StringVar()
route= tk.StringVar()
idprefix = tk.StringVar()
routeno = tk.StringVar()
mpsuffix = tk.StringVar()

county_list=['AA']
directionList=()


def county_callback(var, indx, mode):
    print ("Traced variable {}".format(selected_county.get()))
    
selected_county.trace_add('write', county_callback)

### Populating the county list based on the data collection year ###

def countyfun(e):
    dataYear=int(selected_data_year.get())
    county_list=[]
    if dataYear == 2019:
        with connection.cursor() as cursor:
            cursor.execute("""\
            select unique county from edw19_base_1000000 where status='Matched' 
            order by county""")
            county_list = cursor.fetchall()
            county_list = [x[0] for x in county_list]
    elif dataYear == 2020:
        with connection.cursor() as cursor:
            cursor.execute("""\
            select unique county from edw20_base_1000000 where status='Matched' 
            order by county""")
            county_list = cursor.fetchall()
            county_list = [x[0] for x in county_list]
    else:
        with connection.cursor() as cursor:
            cursor.execute("""\
            select unique county from edw21_base_1000000 where status='Matched' 
            order by county""")
            county_list = cursor.fetchall()
            county_list = [x[0] for x in county_list]
    county_input.config(value=county_list)
    print(dataYear)
    print(county_list)

#def countyList():
#    return county_list

### Populating the route list based on the data year and county

def routefun(e):
    dataYear=int(selected_data_year.get())
    county=county_input.get()
    if dataYear == 2019:
        with connection.cursor() as cursor:
            query = "select unique ID_PREFIX||' '||ID_ROUTE_NO||' '||MP_SUFFIX from edw19_base_1000000 where status='Matched' and county= :county order by ID_PREFIX||' '||ID_ROUTE_NO||' '||MP_SUFFIX"
            cursor.execute(query, county= county)
            #from pandas import DataFrame
            route_list = cursor.fetchall()
            #df1.columns = [x[0] for x in cursor.description]
            #print("I got %d lines " % len(df1))
            route_list = [x[0].strip() for x in route_list]
    elif year_input.get() == 2020:
        with connection.cursor() as cursor:
            cursor.execute("""\
            select unique ID_PREFIX||' '||ID_ROUTE_NO||' '||MP_SUFFIX from edw20_base_1000000 where status='Matched' and county= :county 
            order by ID_PREFIX||' '||ID_ROUTE_NO||' '||MP_SUFFIX""", county=county)
            #from pandas import DataFrame
            route_list = cursor.fetchall()
            #df1.columns = [x[0] for x in cursor.description]
            #print("I got %d lines " % len(df1))
            route_list = [x[0].strip() for x in route_list]
    else:
        with connection.cursor() as cursor:
            cursor.execute("""\
            select unique ID_PREFIX||' '||ID_ROUTE_NO||' '||MP_SUFFIX from edw21_base_1000000 where status='Matched' and county= :county 
            order by ID_PREFIX||' '||ID_ROUTE_NO||' '||MP_SUFFIX""", county=county)
            #from pandas import DataFrame
            route_list = cursor.fetchall()
            #df1.columns = [x[0] for x in cursor.description]
            #print("I got %d lines " % len(df1))
            route_list = [x[0].strip() for x in route_list]
    route_input.config(value=route_list)
    print(county)
    print(route_list)
    
### Populating direction based on the selected route ###

def dirfun(e):
    dataYear=int(selected_data_year.get())
    county=county_input.get()
    route=route_input.get()
    if dataYear == 2019:
        with connection.cursor() as cursor:
            query = "select unique direction from edw19_base_1000000 where status='Matched' and county= :county and trim(ID_PREFIX||' '||ID_ROUTE_NO||' '||MP_SUFFIX)= :route order by direction"
            cursor.execute(query, county=county, route=route)
            #from pandas import DataFrame
            dir_list = cursor.fetchall()
            #df1.columns = [x[0] for x in cursor.description]
            #print("I got %d lines " % len(df1))
            dir_list = [x[0].strip() for x in dir_list]
    elif year_input.get() == 2020:
        with connection.cursor() as cursor:
            query = "select unique direction from edw19_base_1000000 where status='Matched' and county= :county and trim(ID_PREFIX||' '||ID_ROUTE_NO||' '||MP_SUFFIX)= :route order by direction"
            cursor.execute(query, county=county, route=route)
            #from pandas import DataFrame
            dir_list = cursor.fetchall()
            #df1.columns = [x[0] for x in cursor.description]
            #print("I got %d lines " % len(df1))
            dir_list = [x[0].strip() for x in dir_list]
    else:
        with connection.cursor() as cursor:
            query = "select unique direction from edw19_base_1000000 where status='Matched' and county= :county and trim(ID_PREFIX||' '||ID_ROUTE_NO||' '||MP_SUFFIX)= :route order by direction"
            cursor.execute(query, county=county, route=route)
            #from pandas import DataFrame
            dir_list = cursor.fetchall()
            #df1.columns = [x[0] for x in cursor.description]
            #print("I got %d lines " % len(df1))
            dir_list = [x[0].strip() for x in dir_list]
    dir_list.append('All')
    dir_input.config(value=dir_list)
    print(dir_list)
#selected_county.trace_add('write', routefun)

### Getting Global Route ID and Sub Route ID ###

def grfun(e):
    dataYear=selected_data_year.get()
    county=county_input.get()
    route=route_input.get()
    direction=dir_input.get()
    if direction=='All':
        with connection.cursor() as cursor:
            query = """select unique global_route_id, sub_route_id from edw19_base_1000000 where status='Matched' and county= :county and 
            trim(ID_PREFIX||' '||ID_ROUTE_NO||' '||MP_SUFFIX)= :route and 
            direction in ('N', 'S', 'E', 'W') order by global_route_id, sub_route_id"""
            cursor.execute(query, county= county, route=route)
            from pandas import DataFrame
            dftest = DataFrame(cursor.fetchall())
            dftest.columns = [x[0] for x in cursor.description]
            print("I got %d lines " % len(dftest))
    else:
        with connection.cursor() as cursor:
            query = """select unique global_route_id, sub_route_id from edw19_base_1000000 where status='Matched' and county= :county and 
            trim(ID_PREFIX||' '||ID_ROUTE_NO||' '||MP_SUFFIX)= :route and 
            direction= :direction order by global_route_id, sub_route_id"""
            cursor.execute(query, county= county, route=route, direction=direction)
            from pandas import DataFrame
            dftest = DataFrame(cursor.fetchall())
            dftest.columns = [x[0] for x in cursor.description]
            print("I got %d lines " % len(dftest))
    global subRouteId
    global globalRouteID
    subRouteId=[int(x) for x in dftest['SUB_ROUTE_ID']]
    globalRouteID=int(dftest['GLOBAL_ROUTE_ID'][0])
    print(f"Global_ROUTE_ID: {globalRouteID} and SUB_ROUTE_ID: {subRouteId}")

### Main function ###

def getResults():
    hmisyear=int(hmis_input.get())
    dataYear=int(selected_data_year.get())
    county=county_input.get()
    bmp=float(bmp_input.get())
    emp=float(emp_input.get())
    #subId=int(subRouteId[0])
    globalId=str(globalRouteID)
    global df
    if len(subRouteId)==1:
        subId=int(subRouteId[0])
        print(f"hmisYear: {hmisyear}, , county: {county}, bmp: {bmp}, emp: {emp}, subId: {subId}, globalId: {globalId}")
        with connection.cursor() as cursor:
            cursor.execute("""\
            select
            Q1.routeid,
            Q1.global_route_id,
            Q1.sub_route_id,
            Q1.INV_BMP_ORG,
            Q1.INV_EMP_ORG,
            round(sum(lanes*(a2-a1))/(INV_EMP_ORG-INV_BMP_ORG),3) as no_of_lanes,
            -- non-wheel-path:
            AVG(nvl(long_le_cw_1_8,0) + nvl(long_ctr_cw_1_8,0) +nvl(long_re_cw_1_8,0)+ nvl(trans_cw_1_8,0)) as CW_1_8,
            AVG(nvl(long_le_cw_1_4,0) + nvl(long_ctr_cw_1_4,0) +nvl(long_re_cw_1_4,0)+ nvl(trans_cw_1_4,0)) as CW_1_4,
            AVG(nvl(long_le_cw_3_8,0) + nvl(long_ctr_cw_3_8,0) +nvl(long_re_cw_3_8,0)+ nvl(trans_cw_3_8,0)) as CW_3_8,
            AVG(nvl(long_le_cw_1_2,0) + nvl(long_ctr_cw_1_2,0) +nvl(long_re_cw_1_2,0)+ nvl(trans_cw_1_2,0)) as CW_1_2,
            AVG(nvl(long_le_cw_3_4,0) + nvl(long_ctr_cw_3_4,0) +nvl(long_re_cw_3_4,0)+ nvl(trans_cw_3_4,0)) as CW_3_4,
            AVG(nvl(long_le_cw_1_IN,0) + nvl(long_ctr_cw_1_IN,0) +nvl(long_re_cw_1_IN,0)+ nvl(trans_cw_1_IN,0)) as CW_1_IN,
            AVG(nvl(long_le_cw_1_1_2,0) + nvl(long_ctr_cw_1_1_2,0) +nvl(long_re_cw_1_1_2,0)+ nvl(trans_cw_1_1_2,0)) as CW_1_1_2,
            AVG(nvl(long_le_cw_2_IN,0) + nvl(long_ctr_cw_2_IN,0) +nvl(long_re_cw_2_IN,0)+ nvl(trans_cw_2_IN,0)) as CW_2_IN,
            AVG(nvl(long_le_cw_3_IN,0) + nvl(long_ctr_cw_3_IN,0) +nvl(long_re_cw_3_IN,0)+ nvl(trans_cw_3_IN,0)) as CW_3_IN,
            AVG(nvl(long_le_cw_OVER3,0) + nvl(long_ctr_cw_OVER3,0) +nvl(long_re_cw_OVER3,0)+ nvl(trans_cw_OVER3,0)) as CW_OVER3,
            --wheel-path:
            AVG(nvl(long_lwp_cw_1_8,0) + nvl(long_rwp_cw_1_8,0)) as WP_CW_1_8,
            AVG(nvl(long_lwp_cw_1_4,0) + nvl(long_rwp_cw_1_4,0)) as WP_CW_1_4,
            AVG(nvl(long_lwp_cw_3_8,0) + nvl(long_rwp_cw_3_8,0)) as WP_CW_3_8,
            AVG(nvl(long_lwp_cw_1_2,0) + nvl(long_rwp_cw_1_2,0)) as WP_CW_1_2,
            AVG(nvl(long_lwp_cw_3_4,0) + nvl(long_rwp_cw_3_4,0)) as WP_CW_3_4,
            AVG(nvl(long_lwp_cw_1_IN,0) + nvl(long_rwp_cw_1_IN,0)) as WP_CW_1_IN,
            AVG(nvl(long_lwp_cw_1_1_2,0) + nvl(long_rwp_cw_1_1_2,0)) as WP_CW_1_1_2,
            AVG(nvl(long_lwp_cw_2_IN,0) + nvl(long_rwp_cw_2_IN,0)) as WP_CW_2_IN,
            AVG(nvl(long_lwp_cw_3_IN,0) + nvl(long_rwp_cw_3_IN,0)) as WP_CW_3_IN,
            AVG(nvl(long_lwp_cw_OVER3,0) + nvl(long_rwp_cw_OVER3,0)) as WP_CW_OVER3,
            -- Extrapolated columns
              -- non-wheel-path
            AVG(nvl(long_le_cw_1_8,0) + nvl(long_ctr_cw_1_8,0) +nvl(long_re_cw_1_8,0)+ nvl(trans_cw_1_8,0))*round(sum(lanes*(a2-a1))/(INV_EMP_ORG-INV_BMP_ORG),3) as ex_CW_1_8,
            AVG(nvl(long_le_cw_1_4,0) + nvl(long_ctr_cw_1_4,0) +nvl(long_re_cw_1_4,0)+ nvl(trans_cw_1_4,0))*round(sum(lanes*(a2-a1))/(INV_EMP_ORG-INV_BMP_ORG),3) as ex_CW_1_4,
            AVG(nvl(long_le_cw_3_8,0) + nvl(long_ctr_cw_3_8,0) +nvl(long_re_cw_3_8,0)+ nvl(trans_cw_3_8,0))*round(sum(lanes*(a2-a1))/(INV_EMP_ORG-INV_BMP_ORG),3) as ex_CW_3_8,
            AVG(nvl(long_le_cw_1_2,0) + nvl(long_ctr_cw_1_2,0) +nvl(long_re_cw_1_2,0)+ nvl(trans_cw_1_2,0))*round(sum(lanes*(a2-a1))/(INV_EMP_ORG-INV_BMP_ORG),3) as ex_CW_1_2,
            AVG(nvl(long_le_cw_3_4,0) + nvl(long_ctr_cw_3_4,0) +nvl(long_re_cw_3_4,0)+ nvl(trans_cw_3_4,0))*round(sum(lanes*(a2-a1))/(INV_EMP_ORG-INV_BMP_ORG),3) as ex_CW_3_4,
            AVG(nvl(long_le_cw_1_IN,0) + nvl(long_ctr_cw_1_IN,0) +nvl(long_re_cw_1_IN,0)+ nvl(trans_cw_1_IN,0))*round(sum(lanes*(a2-a1))/(INV_EMP_ORG-INV_BMP_ORG),3) as ex_CW_1_IN,
            AVG(nvl(long_le_cw_1_1_2,0) + nvl(long_ctr_cw_1_1_2,0) +nvl(long_re_cw_1_1_2,0)+ nvl(trans_cw_1_1_2,0))*round(sum(lanes*(a2-a1))/(INV_EMP_ORG-INV_BMP_ORG),3) as ex_CW_1_1_2,
            AVG(nvl(long_le_cw_2_IN,0) + nvl(long_ctr_cw_2_IN,0) +nvl(long_re_cw_2_IN,0)+ nvl(trans_cw_2_IN,0))*round(sum(lanes*(a2-a1))/(INV_EMP_ORG-INV_BMP_ORG),3) as ex_CW_2_IN,
            AVG(nvl(long_le_cw_3_IN,0) + nvl(long_ctr_cw_3_IN,0) +nvl(long_re_cw_3_IN,0)+ nvl(trans_cw_3_IN,0))*round(sum(lanes*(a2-a1))/(INV_EMP_ORG-INV_BMP_ORG),3) as ex_CW_3_IN,
            AVG(nvl(long_le_cw_OVER3,0) + nvl(long_ctr_cw_OVER3,0) +nvl(long_re_cw_OVER3,0)+ nvl(trans_cw_OVER3,0))*round(sum(lanes*(a2-a1))/(INV_EMP_ORG-INV_BMP_ORG),3) as ex_CW_OVER3,
              -- wheel-path
            AVG(nvl(long_lwp_cw_1_8,0) + nvl(long_rwp_cw_1_8,0))*round(sum(lanes*(a2-a1))/(INV_EMP_ORG-INV_BMP_ORG),3) as ex_WP_CW_1_8,
            AVG(nvl(long_lwp_cw_1_4,0) + nvl(long_rwp_cw_1_4,0))*round(sum(lanes*(a2-a1))/(INV_EMP_ORG-INV_BMP_ORG),3) as ex_WP_CW_1_4,
            AVG(nvl(long_lwp_cw_3_8,0) + nvl(long_rwp_cw_3_8,0))*round(sum(lanes*(a2-a1))/(INV_EMP_ORG-INV_BMP_ORG),3) as ex_WP_CW_3_8,
            AVG(nvl(long_lwp_cw_1_2,0) + nvl(long_rwp_cw_1_2,0))*round(sum(lanes*(a2-a1))/(INV_EMP_ORG-INV_BMP_ORG),3) as ex_WP_CW_1_2,
            AVG(nvl(long_lwp_cw_3_4,0) + nvl(long_rwp_cw_3_4,0))*round(sum(lanes*(a2-a1))/(INV_EMP_ORG-INV_BMP_ORG),3) as ex_WP_CW_3_4,
            AVG(nvl(long_lwp_cw_1_IN,0) + nvl(long_rwp_cw_1_IN,0))*round(sum(lanes*(a2-a1))/(INV_EMP_ORG-INV_BMP_ORG),3) as ex_WP_CW_1_IN,
            AVG(nvl(long_lwp_cw_1_1_2,0) + nvl(long_rwp_cw_1_1_2,0))*round(sum(lanes*(a2-a1))/(INV_EMP_ORG-INV_BMP_ORG),3) as ex_WP_CW_1_1_2,
            AVG(nvl(long_lwp_cw_2_IN,0) + nvl(long_rwp_cw_2_IN,0))*round(sum(lanes*(a2-a1))/(INV_EMP_ORG-INV_BMP_ORG),3) as ex_WP_CW_2_IN,
            AVG(nvl(long_lwp_cw_3_IN,0) + nvl(long_rwp_cw_3_IN,0))*round(sum(lanes*(a2-a1))/(INV_EMP_ORG-INV_BMP_ORG),3) as ex_WP_CW_3_IN,
            AVG(nvl(long_lwp_cw_OVER3,0) + nvl(long_rwp_cw_OVER3,0))*round(sum(lanes*(a2-a1))/(INV_EMP_ORG-INV_BMP_ORG),3) as ex_WP_CW_OVER3
              
            from
                    (select unique
                     CASE WHEN 
                    crack.sub_route_id=1 then 
                    Nvl(RT_THRU_LA,0) + 
                    nvl(RT_IN_AUX_NUMIA,0) +
                    nvl(RT_OUT_AUX_NUMIA,0) +
                    (RT_OUT_SHLD_WD+RT_IN_SHLD_WD)/(12) +
                     (case when hmis.facility_type in (1, 2) and hmis.median_ty in (4, 5) then 0.5*MEDIAN_WD/12 else 0 end)
                    else
                    nvl(LT_THRU_LA,0) + 
                    nvl(lt_out_aux_numia,0) + 
                    nvl(LT_IN_AUX_NUMIA,0) + 
                    (LT_OUT_SHLD_WD+LT_IN_SHLD_WD)/(12) +
                     (case when hmis.facility_type in (1, 2) and hmis.median_ty in (4, 5) then 0.5*MEDIAN_WD/12 else 0 end)
                    end lanes,
                    0.5*(ID_MP+crack.INV_BMP_ORG)+0.5*abs(ID_MP-crack.INV_BMP_ORG) a1,
                    0.5*(ID_MP + SECTION_LENGTH+crack.INV_EMP_ORG) - 0.5*abs(ID_MP + SECTION_LENGTH-crack.INV_EMP_ORG) a2 ,
                   crack.global_route_id, crack.sub_route_id, crack.INV_BMP_ORG, crack.INV_EMP_ORG,
                    crack.long_lwp_cw_1_8, crack.long_rwp_cw_1_8, crack.long_le_cw_1_8, crack.long_re_cw_1_8, crack.long_ctr_cw_1_8, crack. trans_cw_1_8,
                    crack.long_lwp_cw_1_4, crack.long_rwp_cw_1_4, crack.long_le_cw_1_4, crack.long_re_cw_1_4, crack.long_ctr_cw_1_4, crack. trans_cw_1_4, 
            crack.long_lwp_cw_3_8, crack.long_rwp_cw_3_8, crack.long_le_cw_3_8, crack.long_re_cw_3_8, crack.long_ctr_cw_3_8, crack. trans_cw_3_8, 
            crack.long_lwp_cw_1_2, crack.long_rwp_cw_1_2, crack.long_le_cw_1_2, crack.long_re_cw_1_2, crack.long_ctr_cw_1_2, crack. trans_cw_1_2, 
            crack.long_lwp_cw_3_4, crack.long_rwp_cw_3_4, crack.long_le_cw_3_4, crack.long_re_cw_3_4, crack.long_ctr_cw_3_4, crack. trans_cw_3_4, 
            crack.long_lwp_cw_1_IN, crack.long_rwp_cw_1_IN, crack.long_le_cw_1_IN, crack.long_re_cw_1_IN, crack.long_ctr_cw_1_IN, crack. trans_cw_1_IN, 
            crack.long_lwp_cw_1_1_2, crack.long_rwp_cw_1_1_2, crack.long_le_cw_1_1_2, crack.long_re_cw_1_1_2, crack.long_ctr_cw_1_1_2, crack. trans_cw_1_1_2, 
            crack.long_lwp_cw_2_IN, crack.long_rwp_cw_2_IN, crack.long_le_cw_2_IN, crack.long_re_cw_2_IN, crack.long_ctr_cw_2_IN, crack. trans_cw_2_IN, 
            crack.long_lwp_cw_3_IN, crack.long_rwp_cw_3_IN, crack.long_le_cw_3_IN, crack.long_re_cw_3_IN, crack.long_ctr_cw_3_IN, crack. trans_cw_3_IN, 
            crack.long_lwp_cw_OVER3, crack.long_rwp_cw_OVER3, crack.long_le_cw_OVER3, crack.long_re_cw_OVER3, crack.long_ctr_cw_OVER3, crack.trans_cw_OVER3,
            HMIS.routeid, hmis.LT_ROADWAY_WD, hmis.RT_ROADWAY_WD
                    from HMIS_Universe_all_years hmis
            -- used right join to pull data from both directions from crack17_base_4 table. HMIS table has only one record for each inventory-direction milepoint (left and right lanes)
                    right join CRACKING_DETAILED crack
                    on 
                    crack.GLOBAL_ROUTE_ID = hmis.GLOBAL_ROUTE_ID
                    where hmis.YEAR = to_number(:hmisyear) and 
                    hmis.GLOBAL_ROUTE_ID= to_number(:globalId)
                  and crack.sub_route_id in to_number(:subId)
               --     hmis.MAIN_LINE NOT IN (5, 6,7, 8, 9) 
                    and crack.STATUS='Matched' 
                    AND crack.INV_EMP_ORG>crack.INV_BMP_ORG
                   AND crack.COLLECT_YEAR= to_number(:dataYear) and crack.COUNTY_ORG= to_char(:county)
                   and crack.INV_BMP_ORG >= to_number(:bmp) and crack.INV_EMP_ORG <= to_number(:emp) and hmis.facility_type in (1,2) and
                    (
                                (ID_MP <=  crack.INV_BMP_ORG  AND ID_MP+Section_length >= crack.INV_EMP_ORG) 
                             OR (ID_MP <=crack.INV_BMP_ORG AND ID_MP+Section_length > crack.INV_BMP_ORG) 
                             OR (ID_MP  <  crack.INV_EMP_ORG   AND ID_MP+Section_length    >=crack.INV_EMP_ORG )
                             OR (ID_MP   > crack.INV_BMP_ORG  AND ID_MP+Section_length   <  crack.INV_EMP_ORG)
                             OR (ID_MP   >= crack.INV_BMP_ORG  AND ID_MP+Section_length  <= crack.INV_EMP_ORG)
                             )  order by a1) Q1
                   group by Q1.routeid, Q1.global_route_id, Q1.sub_route_id, Q1.INV_BMP_ORG, Q1.INV_EMP_ORG
            order by Q1.sub_route_id, Q1.INV_BMP_ORG""", hmisyear=hmisyear, globalId=globalId, subId=subId, dataYear=dataYear, county=county, bmp=bmp, emp=emp)
            from pandas import DataFrame
            df = DataFrame(cursor.fetchall())
            df.columns = [x[0] for x in cursor.description]
            #print(df)
            print("I got %d lines " % len(df)) 
    else:
        subId=subRouteId
        print(f"hmisYear: {hmisyear}, , county: {county}, bmp: {bmp}, emp: {emp}, subId: {subId}, globalId: {globalId}")
        with connection.cursor() as cursor:
            cursor.execute("""\
            select
            Q1.routeid,
            Q1.global_route_id,
            Q1.sub_route_id,
            Q1.INV_BMP_ORG,
            Q1.INV_EMP_ORG,
            round(sum(lanes*(a2-a1))/(INV_EMP_ORG-INV_BMP_ORG),3) as no_of_lanes,
            -- non-wheel-path:
            AVG(nvl(long_le_cw_1_8,0) + nvl(long_ctr_cw_1_8,0) +nvl(long_re_cw_1_8,0)+ nvl(trans_cw_1_8,0)) as CW_1_8,
            AVG(nvl(long_le_cw_1_4,0) + nvl(long_ctr_cw_1_4,0) +nvl(long_re_cw_1_4,0)+ nvl(trans_cw_1_4,0)) as CW_1_4,
            AVG(nvl(long_le_cw_3_8,0) + nvl(long_ctr_cw_3_8,0) +nvl(long_re_cw_3_8,0)+ nvl(trans_cw_3_8,0)) as CW_3_8,
            AVG(nvl(long_le_cw_1_2,0) + nvl(long_ctr_cw_1_2,0) +nvl(long_re_cw_1_2,0)+ nvl(trans_cw_1_2,0)) as CW_1_2,
            AVG(nvl(long_le_cw_3_4,0) + nvl(long_ctr_cw_3_4,0) +nvl(long_re_cw_3_4,0)+ nvl(trans_cw_3_4,0)) as CW_3_4,
            AVG(nvl(long_le_cw_1_IN,0) + nvl(long_ctr_cw_1_IN,0) +nvl(long_re_cw_1_IN,0)+ nvl(trans_cw_1_IN,0)) as CW_1_IN,
            AVG(nvl(long_le_cw_1_1_2,0) + nvl(long_ctr_cw_1_1_2,0) +nvl(long_re_cw_1_1_2,0)+ nvl(trans_cw_1_1_2,0)) as CW_1_1_2,
            AVG(nvl(long_le_cw_2_IN,0) + nvl(long_ctr_cw_2_IN,0) +nvl(long_re_cw_2_IN,0)+ nvl(trans_cw_2_IN,0)) as CW_2_IN,
            AVG(nvl(long_le_cw_3_IN,0) + nvl(long_ctr_cw_3_IN,0) +nvl(long_re_cw_3_IN,0)+ nvl(trans_cw_3_IN,0)) as CW_3_IN,
            AVG(nvl(long_le_cw_OVER3,0) + nvl(long_ctr_cw_OVER3,0) +nvl(long_re_cw_OVER3,0)+ nvl(trans_cw_OVER3,0)) as CW_OVER3,
            --wheel-path:
            AVG(nvl(long_lwp_cw_1_8,0) + nvl(long_rwp_cw_1_8,0)) as WP_CW_1_8,
            AVG(nvl(long_lwp_cw_1_4,0) + nvl(long_rwp_cw_1_4,0)) as WP_CW_1_4,
            AVG(nvl(long_lwp_cw_3_8,0) + nvl(long_rwp_cw_3_8,0)) as WP_CW_3_8,
            AVG(nvl(long_lwp_cw_1_2,0) + nvl(long_rwp_cw_1_2,0)) as WP_CW_1_2,
            AVG(nvl(long_lwp_cw_3_4,0) + nvl(long_rwp_cw_3_4,0)) as WP_CW_3_4,
            AVG(nvl(long_lwp_cw_1_IN,0) + nvl(long_rwp_cw_1_IN,0)) as WP_CW_1_IN,
            AVG(nvl(long_lwp_cw_1_1_2,0) + nvl(long_rwp_cw_1_1_2,0)) as WP_CW_1_1_2,
            AVG(nvl(long_lwp_cw_2_IN,0) + nvl(long_rwp_cw_2_IN,0)) as WP_CW_2_IN,
            AVG(nvl(long_lwp_cw_3_IN,0) + nvl(long_rwp_cw_3_IN,0)) as WP_CW_3_IN,
            AVG(nvl(long_lwp_cw_OVER3,0) + nvl(long_rwp_cw_OVER3,0)) as WP_CW_OVER3,
            -- Extrapolated columns
              -- non-wheel-path
            AVG(nvl(long_le_cw_1_8,0) + nvl(long_ctr_cw_1_8,0) +nvl(long_re_cw_1_8,0)+ nvl(trans_cw_1_8,0))*round(sum(lanes*(a2-a1))/(INV_EMP_ORG-INV_BMP_ORG),3) as ex_CW_1_8,
            AVG(nvl(long_le_cw_1_4,0) + nvl(long_ctr_cw_1_4,0) +nvl(long_re_cw_1_4,0)+ nvl(trans_cw_1_4,0))*round(sum(lanes*(a2-a1))/(INV_EMP_ORG-INV_BMP_ORG),3) as ex_CW_1_4,
            AVG(nvl(long_le_cw_3_8,0) + nvl(long_ctr_cw_3_8,0) +nvl(long_re_cw_3_8,0)+ nvl(trans_cw_3_8,0))*round(sum(lanes*(a2-a1))/(INV_EMP_ORG-INV_BMP_ORG),3) as ex_CW_3_8,
            AVG(nvl(long_le_cw_1_2,0) + nvl(long_ctr_cw_1_2,0) +nvl(long_re_cw_1_2,0)+ nvl(trans_cw_1_2,0))*round(sum(lanes*(a2-a1))/(INV_EMP_ORG-INV_BMP_ORG),3) as ex_CW_1_2,
            AVG(nvl(long_le_cw_3_4,0) + nvl(long_ctr_cw_3_4,0) +nvl(long_re_cw_3_4,0)+ nvl(trans_cw_3_4,0))*round(sum(lanes*(a2-a1))/(INV_EMP_ORG-INV_BMP_ORG),3) as ex_CW_3_4,
            AVG(nvl(long_le_cw_1_IN,0) + nvl(long_ctr_cw_1_IN,0) +nvl(long_re_cw_1_IN,0)+ nvl(trans_cw_1_IN,0))*round(sum(lanes*(a2-a1))/(INV_EMP_ORG-INV_BMP_ORG),3) as ex_CW_1_IN,
            AVG(nvl(long_le_cw_1_1_2,0) + nvl(long_ctr_cw_1_1_2,0) +nvl(long_re_cw_1_1_2,0)+ nvl(trans_cw_1_1_2,0))*round(sum(lanes*(a2-a1))/(INV_EMP_ORG-INV_BMP_ORG),3) as ex_CW_1_1_2,
            AVG(nvl(long_le_cw_2_IN,0) + nvl(long_ctr_cw_2_IN,0) +nvl(long_re_cw_2_IN,0)+ nvl(trans_cw_2_IN,0))*round(sum(lanes*(a2-a1))/(INV_EMP_ORG-INV_BMP_ORG),3) as ex_CW_2_IN,
            AVG(nvl(long_le_cw_3_IN,0) + nvl(long_ctr_cw_3_IN,0) +nvl(long_re_cw_3_IN,0)+ nvl(trans_cw_3_IN,0))*round(sum(lanes*(a2-a1))/(INV_EMP_ORG-INV_BMP_ORG),3) as ex_CW_3_IN,
            AVG(nvl(long_le_cw_OVER3,0) + nvl(long_ctr_cw_OVER3,0) +nvl(long_re_cw_OVER3,0)+ nvl(trans_cw_OVER3,0))*round(sum(lanes*(a2-a1))/(INV_EMP_ORG-INV_BMP_ORG),3) as ex_CW_OVER3,
              -- wheel-path
            AVG(nvl(long_lwp_cw_1_8,0) + nvl(long_rwp_cw_1_8,0))*round(sum(lanes*(a2-a1))/(INV_EMP_ORG-INV_BMP_ORG),3) as ex_WP_CW_1_8,
            AVG(nvl(long_lwp_cw_1_4,0) + nvl(long_rwp_cw_1_4,0))*round(sum(lanes*(a2-a1))/(INV_EMP_ORG-INV_BMP_ORG),3) as ex_WP_CW_1_4,
            AVG(nvl(long_lwp_cw_3_8,0) + nvl(long_rwp_cw_3_8,0))*round(sum(lanes*(a2-a1))/(INV_EMP_ORG-INV_BMP_ORG),3) as ex_WP_CW_3_8,
            AVG(nvl(long_lwp_cw_1_2,0) + nvl(long_rwp_cw_1_2,0))*round(sum(lanes*(a2-a1))/(INV_EMP_ORG-INV_BMP_ORG),3) as ex_WP_CW_1_2,
            AVG(nvl(long_lwp_cw_3_4,0) + nvl(long_rwp_cw_3_4,0))*round(sum(lanes*(a2-a1))/(INV_EMP_ORG-INV_BMP_ORG),3) as ex_WP_CW_3_4,
            AVG(nvl(long_lwp_cw_1_IN,0) + nvl(long_rwp_cw_1_IN,0))*round(sum(lanes*(a2-a1))/(INV_EMP_ORG-INV_BMP_ORG),3) as ex_WP_CW_1_IN,
            AVG(nvl(long_lwp_cw_1_1_2,0) + nvl(long_rwp_cw_1_1_2,0))*round(sum(lanes*(a2-a1))/(INV_EMP_ORG-INV_BMP_ORG),3) as ex_WP_CW_1_1_2,
            AVG(nvl(long_lwp_cw_2_IN,0) + nvl(long_rwp_cw_2_IN,0))*round(sum(lanes*(a2-a1))/(INV_EMP_ORG-INV_BMP_ORG),3) as ex_WP_CW_2_IN,
            AVG(nvl(long_lwp_cw_3_IN,0) + nvl(long_rwp_cw_3_IN,0))*round(sum(lanes*(a2-a1))/(INV_EMP_ORG-INV_BMP_ORG),3) as ex_WP_CW_3_IN,
            AVG(nvl(long_lwp_cw_OVER3,0) + nvl(long_rwp_cw_OVER3,0))*round(sum(lanes*(a2-a1))/(INV_EMP_ORG-INV_BMP_ORG),3) as ex_WP_CW_OVER3
              
            from
                    (select unique
                     CASE WHEN 
                    crack.sub_route_id=1 then 
                    Nvl(RT_THRU_LA,0) + 
                    nvl(RT_IN_AUX_NUMIA,0) +
                    nvl(RT_OUT_AUX_NUMIA,0) +
                    (RT_OUT_SHLD_WD+RT_IN_SHLD_WD)/(12) +
                     (case when hmis.facility_type in (1, 2) and hmis.median_ty in (4, 5) then 0.5*MEDIAN_WD/12 else 0 end)
                    else
                    nvl(LT_THRU_LA,0) + 
                    nvl(lt_out_aux_numia,0) + 
                    nvl(LT_IN_AUX_NUMIA,0) + 
                    (LT_OUT_SHLD_WD+LT_IN_SHLD_WD)/(12) +
                     (case when hmis.facility_type in (1, 2) and hmis.median_ty in (4, 5) then 0.5*MEDIAN_WD/12 else 0 end)
                    end lanes,
                    0.5*(ID_MP+crack.INV_BMP_ORG)+0.5*abs(ID_MP-crack.INV_BMP_ORG) a1,
                    0.5*(ID_MP + SECTION_LENGTH+crack.INV_EMP_ORG) - 0.5*abs(ID_MP + SECTION_LENGTH-crack.INV_EMP_ORG) a2 ,
                   crack.global_route_id, crack.sub_route_id, crack.INV_BMP_ORG, crack.INV_EMP_ORG,
                    crack.long_lwp_cw_1_8, crack.long_rwp_cw_1_8, crack.long_le_cw_1_8, crack.long_re_cw_1_8, crack.long_ctr_cw_1_8, crack. trans_cw_1_8,
                    crack.long_lwp_cw_1_4, crack.long_rwp_cw_1_4, crack.long_le_cw_1_4, crack.long_re_cw_1_4, crack.long_ctr_cw_1_4, crack. trans_cw_1_4, 
            crack.long_lwp_cw_3_8, crack.long_rwp_cw_3_8, crack.long_le_cw_3_8, crack.long_re_cw_3_8, crack.long_ctr_cw_3_8, crack. trans_cw_3_8, 
            crack.long_lwp_cw_1_2, crack.long_rwp_cw_1_2, crack.long_le_cw_1_2, crack.long_re_cw_1_2, crack.long_ctr_cw_1_2, crack. trans_cw_1_2, 
            crack.long_lwp_cw_3_4, crack.long_rwp_cw_3_4, crack.long_le_cw_3_4, crack.long_re_cw_3_4, crack.long_ctr_cw_3_4, crack. trans_cw_3_4, 
            crack.long_lwp_cw_1_IN, crack.long_rwp_cw_1_IN, crack.long_le_cw_1_IN, crack.long_re_cw_1_IN, crack.long_ctr_cw_1_IN, crack. trans_cw_1_IN, 
            crack.long_lwp_cw_1_1_2, crack.long_rwp_cw_1_1_2, crack.long_le_cw_1_1_2, crack.long_re_cw_1_1_2, crack.long_ctr_cw_1_1_2, crack. trans_cw_1_1_2, 
            crack.long_lwp_cw_2_IN, crack.long_rwp_cw_2_IN, crack.long_le_cw_2_IN, crack.long_re_cw_2_IN, crack.long_ctr_cw_2_IN, crack. trans_cw_2_IN, 
            crack.long_lwp_cw_3_IN, crack.long_rwp_cw_3_IN, crack.long_le_cw_3_IN, crack.long_re_cw_3_IN, crack.long_ctr_cw_3_IN, crack. trans_cw_3_IN, 
            crack.long_lwp_cw_OVER3, crack.long_rwp_cw_OVER3, crack.long_le_cw_OVER3, crack.long_re_cw_OVER3, crack.long_ctr_cw_OVER3, crack.trans_cw_OVER3,
            HMIS.routeid, hmis.LT_ROADWAY_WD, hmis.RT_ROADWAY_WD
                    from HMIS_Universe_all_years hmis
            -- used right join to pull data from both directions from crack17_base_4 table. HMIS table has only one record for each inventory-direction milepoint (left and right lanes)
                    right join CRACKING_DETAILED crack
                    on 
                    crack.GLOBAL_ROUTE_ID = hmis.GLOBAL_ROUTE_ID
                    where hmis.YEAR = to_number(:hmisyear) and 
                    hmis.GLOBAL_ROUTE_ID= to_number(:globalId)
                  and crack.sub_route_id in (1,2)
               --     hmis.MAIN_LINE NOT IN (5, 6,7, 8, 9) 
                    and crack.STATUS='Matched' 
                    AND crack.INV_EMP_ORG>crack.INV_BMP_ORG
                   AND crack.COLLECT_YEAR= to_number(:dataYear) and crack.COUNTY_ORG= to_char(:county)
                   and crack.INV_BMP_ORG >= to_number(:bmp) and crack.INV_EMP_ORG <= to_number(:emp) and hmis.facility_type in (1,2) and
                    (
                                (ID_MP <=  crack.INV_BMP_ORG  AND ID_MP+Section_length >= crack.INV_EMP_ORG) 
                             OR (ID_MP <=crack.INV_BMP_ORG AND ID_MP+Section_length > crack.INV_BMP_ORG) 
                             OR (ID_MP  <  crack.INV_EMP_ORG   AND ID_MP+Section_length    >=crack.INV_EMP_ORG )
                             OR (ID_MP   > crack.INV_BMP_ORG  AND ID_MP+Section_length   <  crack.INV_EMP_ORG)
                             OR (ID_MP   >= crack.INV_BMP_ORG  AND ID_MP+Section_length  <= crack.INV_EMP_ORG)
                             )  order by a1) Q1
                   group by Q1.routeid, Q1.global_route_id, Q1.sub_route_id, Q1.INV_BMP_ORG, Q1.INV_EMP_ORG
            order by Q1.sub_route_id, Q1.INV_BMP_ORG""", hmisyear=hmisyear, globalId=globalId, dataYear=dataYear, county=county, bmp=bmp, emp=emp)
            from pandas import DataFrame
            df = DataFrame(cursor.fetchall())
            df.columns = [x[0] for x in cursor.description]
            #print(df)
            print("I got %d lines " % len(df)) 
    # non wheelpath cracks
    cw_1_8 = round(df['EX_CW_1_8'].sum(),3)
    cw_1_4 = round(df['EX_CW_1_4'].sum(),3)
    cw_3_8 = round(df['EX_CW_3_8'].sum(),3)
    cw_1_2 = round(df['EX_CW_1_2'].sum(),3)
    cw_3_4 = round(df['EX_CW_3_4'].sum(),3)
    cw_1_in = round(df['EX_CW_1_IN'].sum(),3)
    cw_1_1_2 = round(df['EX_CW_1_1_2'].sum(),3)
    cw_2_in = round(df['EX_CW_2_IN'].sum(),3)
    cw_3_in = round(df['EX_CW_3_IN'].sum(),3)
    cw_over3 = round(df['EX_CW_OVER3'].sum(),3)
    
    global nwp
    nwp = [cw_1_8, cw_1_4, cw_3_8, cw_1_2, cw_3_4, cw_1_in, cw_1_1_2, cw_2_in, cw_3_in, cw_over3]

    # wheelpath cracks
    wp_cw_1_8 = df['EX_WP_CW_1_8'].sum()
    wp_cw_1_4 = df['EX_WP_CW_1_4'].sum()
    wp_cw_3_8 = df['EX_WP_CW_3_8'].sum()
    wp_cw_1_2 = df['EX_WP_CW_1_2'].sum()
    wp_cw_3_4 = df['EX_WP_CW_3_4'].sum()
    wp_cw_1_in = df['EX_WP_CW_1_IN'].sum()
    wp_cw_1_1_2 = df['EX_WP_CW_1_1_2'].sum()
    wp_cw_2_in = df['EX_WP_CW_2_IN'].sum()
    wp_cw_3_in = df['EX_WP_CW_3_IN'].sum()
    wp_cw_over3 = df['EX_WP_CW_OVER3'].sum()
    
    global wp
    wp = [wp_cw_1_8, wp_cw_1_4, wp_cw_3_8, wp_cw_1_2, wp_cw_3_4, wp_cw_1_in, wp_cw_1_1_2, wp_cw_2_in, wp_cw_3_in, wp_cw_over3]
    
    
    #scrollbar
    result_scroll = ttk.Scrollbar(result)
    result_scroll.pack(side='right', fill='y')

    result_scroll = ttk.Scrollbar(result,orient='horizontal')
    result_scroll.pack(side= 'bottom',fill='x')

    set = ttk.Treeview(result,yscrollcommand=result_scroll.set, xscrollcommand =result_scroll.set)
    set.pack()

    set['columns']= ('crack_location','cw_1_8', 'cw_1_4', 'cw_3_8', 'cw_1_2', 'cw_3_4', 'cw_1_in', 'cw_1_1_2', 'cw_2_in', 'cw_3_in', 'cw_over3')
    set.column("#0", width=0,  stretch='yes')
    set.column("crack_location",anchor='center', width=120)
    set.column("cw_1_8",anchor='center', width=90)
    set.column("cw_1_4",anchor='center', width=90)
    set.column("cw_3_8",anchor='center', width=90)
    set.column("cw_1_2",anchor='center', width=90)
    set.column("cw_3_4",anchor='center', width=90)
    set.column("cw_1_in",anchor='center', width=90)
    set.column("cw_1_1_2",anchor='center', width=90)
    set.column("cw_2_in",anchor='center', width=90)
    set.column("cw_3_in",anchor='center', width=90)
    set.column("cw_over3",anchor='center', width=100)

    set.heading("#0",text="",anchor='center')
    set.heading("crack_location",text="Crack Location",anchor='center')
    set.heading("cw_1_8",text="cw_1_8",anchor='center')
    set.heading("cw_1_4",text="cw_1_4",anchor='center')
    set.heading("cw_3_8",text="cw_3_8",anchor='center')
    set.heading("cw_1_2",text="cw_1_2",anchor='center')
    set.heading("cw_3_4",text="cw_3_4",anchor='center')
    set.heading("cw_1_in",text="cw_1_in",anchor='center')
    set.heading("cw_1_1_2",text="cw_1_1_2",anchor='center')
    set.heading("cw_2_in",text="cw_2_in",anchor='center')
    set.heading("cw_3_in",text="cw_3_in",anchor='center')
    set.heading("cw_over3",text="cw_over3",anchor='center')

    set.insert(parent='',index='end',iid=0,text='',
           values=('Non Wheel Path', round(cw_1_8,3), round(cw_1_4,3), round(cw_3_8,3), round(cw_1_2,3), 
                   round(cw_3_4,3), round(cw_1_in,3), round(cw_1_1_2,3), round(cw_2_in,3), round(cw_3_in,3), round(cw_over3,3)))

    #### SCROLLBAR 2
    result_scroll2 = ttk.Scrollbar(result)
    result_scroll2.pack(side='right', fill='y')

    result_scroll2 = ttk.Scrollbar(result,orient='horizontal')
    result_scroll2.pack(side= 'bottom',fill='x')

    set = ttk.Treeview(result,yscrollcommand=result_scroll2.set, xscrollcommand =result_scroll2.set)
    set.pack()

    set['columns']= ('crack_location','wp_cw_1_8', 'wp_cw_1_4', 'wp_cw_3_8', 'wp_cw_1_2', 'wp_cw_3_4', 'wp_cw_1_in', 'wp_cw_1_1_2', 'wp_cw_2_in', 'wp_cw_3_in', 'wp_cw_over3')
    set.column("#0", width=0,  stretch='yes')
    set.column("crack_location",anchor='center', width=120)
    set.column("wp_cw_1_8",anchor='center', width=90)
    set.column("wp_cw_1_4",anchor='center', width=90)
    set.column("wp_cw_3_8",anchor='center', width=90)
    set.column("wp_cw_1_2",anchor='center', width=90)
    set.column("wp_cw_3_4",anchor='center', width=90)
    set.column("wp_cw_1_in",anchor='center', width=90)
    set.column("wp_cw_1_1_2",anchor='center', width=90)
    set.column("wp_cw_2_in",anchor='center', width=90)
    set.column("wp_cw_3_in",anchor='center', width=90)
    set.column("wp_cw_over3",anchor='center', width=100)

    set.heading("#0",text="",anchor='center')
    set.heading("crack_location",text="Crack Location",anchor='center')
    set.heading("wp_cw_1_8",text="cw_1_8",anchor='center')
    set.heading("wp_cw_1_4",text="cw_1_4",anchor='center')
    set.heading("wp_cw_3_8",text="cw_3_8",anchor='center')
    set.heading("wp_cw_1_2",text="cw_1_2",anchor='center')
    set.heading("wp_cw_3_4",text="cw_3_4",anchor='center')
    set.heading("wp_cw_1_in",text="cw_1_in",anchor='center')
    set.heading("wp_cw_1_1_2",text="cw_1_1_2",anchor='center')
    set.heading("wp_cw_2_in",text="cw_2_in",anchor='center')
    set.heading("wp_cw_3_in",text="cw_3_in",anchor='center')
    set.heading("wp_cw_over3",text="cw_over3",anchor='center')

    set.insert(parent='',index='end',iid=0,text='',
           values=('Wheel Path', round(wp_cw_1_8,3), round(wp_cw_1_4,3), round(wp_cw_3_8,3), round(wp_cw_1_2,3), 
                   round(wp_cw_3_4,3), round(wp_cw_1_in,3), round(wp_cw_1_1_2,3), round(wp_cw_2_in,3), round(wp_cw_3_in,3), round(wp_cw_over3,3)))
    
    
def extract():
    dataYear=int(selected_data_year.get())
    county=county_input.get()
    route=route_input.get()
    direction=dir_input.get()
    bmp=float(bmp_input.get())
    emp=float(emp_input.get())
    #subId=int(subRouteId[0])
        
    myworkbook=openpyxl.load_workbook('Cracking-Quantity.xlsx')
    ws = myworkbook["mainsheet"]
    ws1 = myworkbook.copy_worksheet(ws)
    ws1.title = f'{county}-{route}-MP-{bmp}-{emp}'
    myworkbook.save('Cracking-Quantity.xlsx')
    #worksheet= myworkbook['mainsheet']
    myworkbook=openpyxl.load_workbook('Cracking-Quantity.xlsx')
    ws = myworkbook[f'{county}-{route}-MP-{bmp}-{emp}']
    nwpcells = ['C9', 'D9', 'E9', 'F9', 'G9', 'H9', 'I9', 'J9', 'K9', 'L9']
    wpcells = ['C31', 'D31', 'E31', 'F31', 'G31', 'H31', 'I31', 'J31', 'K31', 'L31']

    ws['L1'] = date.today().strftime("%d/%m/%Y")

    ws['A1']= f'{county} {route} - MP {bmp} to {emp} (Crack Sealing Quantity - Non Wheel Path)'
    ws['A23']= f'{county} {route} - MP {bmp} to {emp} (Crack Sealing Quantity - Wheel Path Only)'
    ws['C7'] = f'{county} {route} - MP: {bmp}-{emp}, {direction}  (Non Wheel Path)'
    ws['C29'] = f'{county} {route} - MP: {bmp}-{emp}, {direction}  (Wheel Path Only)'

    # filling in the crack length data for each crack width

    for i in range(0 , len(nwpcells)):
        ws[nwpcells[i]]= nwp[i]
        ws[wpcells[i]]= wp[i]

    # filling in the sum of crack length data 
     # Non Wheel Path
    ws['H10']= sum(nwp[:6])
    ws['L10']= sum(nwp[6:])

     # Wheel Path
    ws['H32']= sum(wp[:6])
    ws['L32']= sum(wp[6:])

    myworkbook.save('Cracking-Quantity.xlsx')
    
##### main frame (input) configuration: #####

hmis_label = ttk.Label(main, text="HMIS YEAR:")
hmis_label.grid(column=0, row=1, sticky="W", padx=5)
hmis_input = ttk.Combobox(main, textvariable=selected_hmis_year)
hmis_input['state'] = 'readonly'
hmis_input['values'] = (2016)
hmis_input.grid(column=0, row=2, sticky="EW", padx=5)
#hmis_input.current(0)

year_label = ttk.Label(main, text="DATA YEAR:")
year_label.grid(column=1, row=1, sticky="W", padx=5)
year_input = ttk.Combobox(main, textvariable=selected_data_year)
year_input['state'] = 'readonly'
#year_input.bind('<<ComboboxSelected>>', hmis_input)
year = datetime.datetime.today().year
year_input['values'] = list(range(year, year - 3, -1))
year_input.grid(column=1, row=2, sticky="EW", padx=5)
year_input.bind('<<ComboboxSelected>>', countyfun)
#year_input.current(0)

county_label = ttk.Label(main, text="COUNTY:")
county_label.grid(column=2, row=1, sticky="W", padx=5)
county_input = ttk.Combobox(main, textvariable=selected_county)
county_input['state'] = 'readonly'
county_input.bind('<<ComboboxSelected>>', routefun)
#county_input['values'] = countyList()
county_input.grid(column=2, row=2, sticky="EW", padx=5)
#county_input.current(0)


route_label = ttk.Label(main, text="ROUTE:")
route_label.grid(column=3, row=1, sticky="W", padx=5)
route_input = ttk.Combobox(main, textvariable=selected_route)
route_input.bind('<<ComboboxSelected>>', dirfun)
#route_input['values'] = route_list
route_input.grid(column=3, row=2, sticky="EW", padx=5)

dir_label = ttk.Label(main, text="DIRECTION:")
dir_label.grid(column=4, row=1, sticky="W", padx=5)
dir_input = ttk.Combobox(main, textvariable=direction)
dir_input['state'] = 'readonly'
#route_input['values'] = route_list
dir_input.grid(column=4, row=2, sticky="EW", padx=5)
dir_input.bind('<<ComboboxSelected>>', grfun)

bmp_label = ttk.Label(main, text="BMP:")
bmp_label.grid(column=0, row=3, columnspan=1, sticky="W", padx=5)
bmp_input = ttk.Entry(main, width=10, textvariable=bmp, font=(None, 12))  # None means "don't change the font".
bmp_input.grid(column=0, row=4, columnspan=1, sticky="EW", padx=5)

emp_label = ttk.Label(main, text="EMP:")
emp_label.grid(column=1, row=3, columnspan=1, sticky="W", padx=5)
emp_input = ttk.Entry(main, width=10, textvariable=emp, font=(None, 12))  # None means "don't change the font".
emp_input.grid(column=1, row=4, columnspan=1, sticky="EW", padx=5)
emp_input.bind('<<ComboboxSelected>>', grfun)

calc_button = ttk.Button(main, text="Enter", command=getResults)
calc_button.grid(column=2, row=4, columnspan=1, sticky="EW", padx=5)

extract_button = ttk.Button(main, text="Extract", command=extract)
extract_button.grid(column=3, row=4, columnspan=1, sticky="EW", padx=5)

closeButton = tk.Button(main, text="Close", width=15, command=root.destroy).grid(row=4, column=5)

##### result frame (output) configuration: #####
"""
cw_1_8_label = ttk.Label(result, text="cw_1_8:")
cw_1_8_label.grid(column=0, row=1, sticky="W", padx=5)
cw_1_8_input = ttk.Label(result, text=0.5)
cw_1_8_input.grid(column=0, row=2, sticky="EW", padx=5)

cw_1_4_label = ttk.Label(result, text="cw_1_8:")
cw_1_4_label.grid(column=1, row=1, sticky="W", padx=5)
cw_1_4_input = ttk.Label(result, text=0.5)
cw_1_4_input.grid(column=1, row=2, sticky="EW", padx=5)
"""
root.mainloop()